"""
A_rho Calculator — Quantum Assembly Theory (QAT)
======================================================
    A_rho = (1/sigma) * { E(M) * d_avg + Ring_cost/N + Cage_cost/N }

    E(M)  = (N + N_rcp + N_ccp) / N
    d_avg = (2/N(N-1)) * sum_{i<j} d_ij
    d_ij  = shortest weighted path, edge weight = -ln(rho_BCP)
"""

import os, re, sys, math
import numpy as np
import networkx as nx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ===========================================================================
# CONFIGURATION
# ===========================================================================

DATA_DIR   = "."
EXCEL_FILE = "Book1.xlsx"
# NAME_MAP = {"Excel molecule name" :"File name"}
NAME_MAP = {
    "Argon"                 : "Argon",
    "Dihydrogen"            : "H2",
    "Dinitrogen"            : "N2",
    "Dioxygen"              : "O2",
    "Water"                 : "H2O",
    "Ammonia"               : "NH3",
    "Benzene"               : "Benzene",
    "Sulfur Hexafluoride"   : "SF6",
    "Ammonia Borane"        : "NH3-BH3",
    "Cyclopropane"          : "C3H6",
    "Adamantane"            : "Adamantane",
    "Cubane"                : "Cubane",
    "L-Alanine"             : "L-Alanine",   
    "D-Glucose"             : "D-Glucose",
    "Caffiene"              : "Caffiene",
    "Penicillin G"          : "Penicillin_G",
    "Cholesterol"           : "Cholesterol",
    "Strychnine"            : "Strychnine",
    "ATP"                   : "ATP",
    "Erythromycin"          : "Erythromycin",
    "Chlorophyll A"         : "Chlorophyll_A",
    "Vitamin B12"           : "Vitamin_B12",
    "Heme B"                : "Heme_B",
    "Morphine"              : "Morphine",
    "Napthalene"            : "Napthalene",
    "Urea"                  : "Urea",
    "Ferrocene"             : "Ferrocene",
    "Nickelocene"           : "Nickelocene",
    "ZSM-5 fragment"        : "ZSM-5",
    "C3H3N3(Triazine)"      : "Triazine",
    "HKUST-1"               : "HKUST1",
    "BaTiO3 (Perovsite)"    : "BaTiO3",     
    "C60 (Fullerene)"       : "C60",
    "In4Sb5 (Zintl Cluster)": "In4Sb5",
    "Mo154+14Li+"           : "Mo154",
    "Alanine_Dipeptide"     : "Alanine_Dipeptide",
    "Glycine"               : "Glycine",
    "Hydrazine"             : "Hydrazine",
    "Aspirin"               : "Aspirin",
    "Cefixime"              : "Cefixime",
    "Nicotine"              : "Nicotine",
    "Serotonin"             : "Serotonin",
    "Formamide"             : "Formamide",
    "B-Alanine"             : "B-Alanine",
    "Sarcosine"             : "Sarcosine",
    "Pyruvic_acid"          : "Pyruvic_acid",
    "Lactic_acid"           : "Lactic_acid",
    "L-Serine"              : "L-Serine",
    "L-Proline"             : "L-Proline",
    "L-Cysteine"            : "L-Cysteine",
    "Citrate"               : "Citrate",
    "Isocitrate"            : "Isocitrate",
    "Adenosine"             : "Adenosine",
    "Ferrioxalate"          : "Ferrioxalate",
    "Diborane"              : "Diborane",
    "Phosphoric_acid"       : "Phosphoric_acid",
    "Hexacyanoferrate"      : "Hexacyanoferrate",
    "Decavanadate"          : "Decavanadate",
    "A-T_base_pair"         : "A-T_base_pair",
    "GC_Dinucleotide"       : "GC_Dinucleotide"
}

COL_MOLECULE  = "Molecule"
COL_NATOM     = "Natom"
COL_NRCP      = "Nrcp"
COL_NCCP      = "Nccp"
COL_RING_COST = "Ring cost"
COL_CAGE_COST = "Cage cost"
COL_SYMMETRY  = "symmetry"
COL_ARHO      = "A_rho"

# ===========================================================================
# FILE DISCOVERY
# ===========================================================================

def find_qtaim_files(base_dir, excel_name):
    """
    Match files using stem from NAME_MAP.

    """
    stem = NAME_MAP.get(excel_name, excel_name)
    stem_lower = stem.lower()
    summary_path = cpprop_path = None

    try:
        all_files = os.listdir(base_dir)
    except FileNotFoundError:
        return None, None

    for fname in all_files:
        fname_lower = fname.lower()
        bare_lower  = re.sub(r'^\d+_', '', fname_lower)

        # Match against full filename OR prefix-stripped filename
        if (fname_lower.startswith(stem_lower + '_') or
                bare_lower.startswith(stem_lower + '_')):
            full = os.path.join(base_dir, fname)
            if fname_lower.endswith('_summary.txt'):
                summary_path = full
            elif fname_lower.endswith('_cpprop.txt'):
                cpprop_path = full

    return summary_path, cpprop_path

# ===========================================================================
# QTAIM PARSING
# ===========================================================================

def parse_bcp_data(summary_path, cpprop_path):
    summary_text = ""
    if summary_path and os.path.exists(summary_path):
        with open(summary_path, 'r', errors='replace') as f:
            summary_text = f.read()

    cpprop_lines = []
    if cpprop_path and os.path.exists(cpprop_path):
        with open(cpprop_path, 'r', errors='replace') as f:
            cpprop_lines = f.read().splitlines()

    # Atom count
    n_atoms = 0
    m = re.search(r'Number of atoms:\s*(\d+)', summary_text)
    if m:
        n_atoms = int(m.group(1))
    if n_atoms == 0:
        n_atoms = len(re.findall(r'\(3,-3\)', summary_text))

    # BCP rho from summary (3,-1) = {...} — skip connectivity blocks
    rho_bcp_list = []
    for m in re.finditer(r'\(3,-1\)\s*=\s*\{(.*?)\}', summary_text, re.DOTALL):
        inner = m.group(1).strip()
        if not inner or '<->' in inner:
            continue
        try:
            vals = [float(x) for x in inner.split(',') if x.strip()]
            if vals:
                rho_bcp_list = vals
                break
        except ValueError:
            continue

    # Connectivity from CPprop
    atom_pairs = []
    for line in cpprop_lines:
        if ("Connected atoms:" in line or
                "Corresponding atoms:" in line or
                "Bond between atom" in line):
            nums = re.findall(r'\b(\d+)\s*\(', line)
            if len(nums) >= 2:
                atom_pairs.append((int(nums[0]), int(nums[1])))

    # Fallback for large molecules (Mo154): parse rho from CPprop per-BCP blocks
    if not rho_bcp_list and atom_pairs:
        rho_from_cpprop = []
        pairs_from_cpprop = []
        in_bcp = False
        current_pair = current_rho = None

        for line in cpprop_lines:
            if 'Type (3,-1)' in line:
                if current_pair and current_rho is not None:
                    pairs_from_cpprop.append(current_pair)
                    rho_from_cpprop.append(current_rho)
                in_bcp = True
                current_pair = current_rho = None
            elif in_bcp:
                if "Connected atoms:" in line or "Corresponding atoms:" in line:
                    nums = re.findall(r'\b(\d+)\s*\(', line)
                    if len(nums) >= 2:
                        current_pair = (int(nums[0]), int(nums[1]))
                elif 'Density of all electrons:' in line:
                    vm = re.search(r':\s*([\d.E+\-]+)', line)
                    if vm:
                        try:
                            current_rho = float(vm.group(1))
                        except ValueError:
                            pass
                elif 'Type (3,' in line and '(3,-1)' not in line:
                    if current_pair and current_rho is not None:
                        pairs_from_cpprop.append(current_pair)
                        rho_from_cpprop.append(current_rho)
                    in_bcp = False
                    current_pair = current_rho = None

        if in_bcp and current_pair and current_rho is not None:
            pairs_from_cpprop.append(current_pair)
            rho_from_cpprop.append(current_rho)

        if rho_from_cpprop:
            rho_bcp_list = rho_from_cpprop
            atom_pairs   = pairs_from_cpprop

    n_bcps = min(len(rho_bcp_list), len(atom_pairs))
    if n_bcps == 0:
        return [], n_atoms

    bonds = [(atom_pairs[i][0], atom_pairs[i][1], rho_bcp_list[i])
             for i in range(n_bcps) if rho_bcp_list[i] > 0]
    return bonds, n_atoms

# ===========================================================================
# GRAPH DISTANCE
# ===========================================================================

def compute_d_avg(bonds, n_atoms):
    if n_atoms <= 1 or not bonds:
        return 0.0

    G = nx.Graph()
    G.add_nodes_from(range(1, n_atoms + 1))
    for a, b, rho in bonds:
        w = -math.log(rho)
        if G.has_edge(a, b):
            if w < G[a][b]['weight']:
                G[a][b]['weight'] = w
        else:
            G.add_edge(a, b, weight=w)

    try:
        lengths = dict(nx.all_pairs_dijkstra_path_length(G, weight='weight'))
    except Exception as e:
        print(f"    [WARNING] Graph distance failed: {e}")
        return 0.0

    raw_sum = 0.0
    nodes = sorted(G.nodes())
    for idx_i, i in enumerate(nodes):
        for j in nodes[idx_i + 1:]:
            d = lengths.get(i, {}).get(j, None)
            if d is not None and d > 0:
                raw_sum += d

    Np = n_atoms * (n_atoms - 1) / 2.0
    return raw_sum / Np if Np > 0 else 0.0

# ===========================================================================
# A_rho FORMULA
# ===========================================================================

def compute_arho(N, N_rcp, N_ccp, ring_cost, cage_cost, sigma, d_avg):
    if N == 0 or sigma == 0:
        return None
    EM = (N + N_rcp + N_ccp) / N
    return (1.0 / sigma) * (EM * d_avg + ring_cost / N + cage_cost / N)

# ===========================================================================
# MAIN
# ===========================================================================

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: '{EXCEL_FILE}' not found.")
        sys.exit(1)

    df = pd.read_excel(EXCEL_FILE)
    df.columns = [c.strip() for c in df.columns]

    for col in [COL_MOLECULE, COL_NATOM, COL_NRCP, COL_NCCP,
                COL_RING_COST, COL_CAGE_COST, COL_SYMMETRY]:
        if col not in df.columns:
            print(f"ERROR: Column '{col}' not found. Available: {list(df.columns)}")
            sys.exit(1)

    arho_values = []
    print(f"\n{'='*70}")
    print(f"  A_rho CALCULATOR — QAT Benchmark")
    print(f"{'='*70}\n")

    for _, row in df.iterrows():
        mol_name  = str(row[COL_MOLECULE]).strip()
        N         = int(row[COL_NATOM])
        N_rcp     = int(row[COL_NRCP])
        N_ccp     = int(row[COL_NCCP])
        ring_cost = float(row[COL_RING_COST])
        cage_cost = float(row[COL_CAGE_COST])
        sigma_raw = row[COL_SYMMETRY]

        if str(sigma_raw).strip().lower() in ('inf', 'infinity', '∞'):
            arho_values.append(0.0)
            print(f"  {mol_name:<30} σ=∞  →  A_rho = 0.0000")
            continue

        sigma = float(sigma_raw)
        summary, cpprop = find_qtaim_files(DATA_DIR, mol_name)

        if summary is None and cpprop is None:
            print(f"  {mol_name:<30} [WARNING] Files not found — d_avg = 0")
            d_avg = 0.0
        else:
            bonds, _ = parse_bcp_data(summary, cpprop)
            if not bonds:
                print(f"  {mol_name:<30} [WARNING] No BCPs parsed — d_avg = 0")
                d_avg = 0.0
            else:
                d_avg = compute_d_avg(bonds, N)

        arho = compute_arho(N, N_rcp, N_ccp, ring_cost, cage_cost, sigma, d_avg)

        if arho is None:
            print(f"  {mol_name:<30} [ERROR] Cannot compute A_rho")
            arho_values.append(None)
        else:
            arho_values.append(round(arho, 4))
            EM = (N + N_rcp + N_ccp) / N
            print(f"  {mol_name:<30} d_avg={d_avg:7.4f}  E(M)={EM:.4f}"
                  f"  σ={sigma:<6.0f}  A_rho = {arho:.4f}")

    # Write to Excel
    df[COL_ARHO] = arho_values
    out_file = "Final_Arho_Results.xlsx"
    df.to_excel(out_file, index=False)

    wb = load_workbook(out_file)
    ws = wb.active
    arho_col = None
    for cell in ws[1]:
        if cell.value == COL_ARHO:
            arho_col = cell.column
            break
    if arho_col:
        hdr = ws.cell(row=1, column=arho_col)
        hdr.font      = Font(bold=True, color="FFFFFF")
        hdr.fill      = PatternFill("solid", start_color="2E4057")
        hdr.alignment = Alignment(horizontal="center")
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=arho_col)
            if c.value is not None:
                c.number_format = "0.0000"
                c.alignment     = Alignment(horizontal="right")
    wb.save(out_file)

    print(f"\n{'='*70}")
    print(f"  Results saved to '{out_file}'")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
